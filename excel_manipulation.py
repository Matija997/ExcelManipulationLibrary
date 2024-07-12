import os
import re
import openpyxl
from openpyxl import load_workbook


class ExcelApi:
    """
    A class to manipulate Excel files.

    Attributes
    ----------
    filename : str
        The path to the Excel file.

    Methods
    -------
    excel_create(data: list = None)
        Creates an Excel file with the provided data.
    excel_delete()
        Deletes the Excel file.
    excel_update(sheetname: str = None, cell: str = None, value=0)
        Updates a specific cell in the specified sheet with a given value.
    excel_open()
        Opens the Excel file and returns the workbook object.
    sheet_create(sheetname: str = None)
        Creates a new sheet in the Excel file.
    sheet_name_change(sheetname: str = None, newname: str = None)
        Renames an existing sheet in the Excel file.
    sheet_delete(sheetname: str = None)
        Deletes a sheet from the Excel file.
    """
    def __init__(self, filename: str = None) -> None:
        """
        Initializes the ExcelApi with the given filename.

        Parameters
        ----------
        filename : str, optional
            The name of the Excel file (default is None).

        Raises
        ------
        ValueError
            If the filename is not provided.
        TypeError
            If the filename is not a string.
        """
        try:

            if filename is None:
                raise ValueError("File name must be a non-empty string")

            if type(filename) is not str:
                raise TypeError("File name must be string")

            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            self.filename = filename

        except ValueError as e:
            print(f"Value error: {e}")

        except TypeError as e:
            print(f"Type error: {e}")

        except Exception as e:
            print(f"An unexcpected error occured: {e}")

    def excel_create(self, data: list = None) -> None:
        """
        Creates an Excel file with the provided data.

        Parameters
        ----------
        data : list, optional
            A list of lists where each sublist represents a row of data
            (default is None).

        Raises
        ------
        ValueError
            If the data is not provided.
        TypeError
            If the data is not a list.
        """
        try:

            if data is None:
                raise ValueError("Data must be a non-empty list")

            if type(data) is not list:
                raise TypeError("Data must be list")

            wb = openpyxl.Workbook()

            sheet = wb.active

            for row_index, row_data in enumerate(data, start=1):
                for col_index, cell_val in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=cell_val)

            wb.save(self.filename)

        except ValueError as e:
            print(f"Value error: {e}")

        except TypeError as e:
            print(f"Type error: {e}")

        except Exception as e:
            print(f"An unexcpected error occured: {e}")

        else:
            print(f"Excel file '{self.filename}' created successfully.")

    def excel_delete(self) -> None:
        """
        Deletes the Excel file.

        Raises
        ------
        OSError
            If there is an error removing the file.
        """
        try:

            if os.path.exists(self.filename):
                os.remove(self.filename)
                print(f"{self.filename} has been successfully deleted.")

            else:
                print(f"{self.filename} does not exist.")

        except OSError as e:
            print(f"Error: {e.strerror}")

        except Exception as e:
            print(f"An unexcepted error occured: {e}")

    def excel_updete(self,
                     sheetname: str = None,
                     cell: str = None,
                     value=0) -> None:
        """
        Updates a specific cell in the specified sheet with a given value.

        Parameters
        ----------
        sheetname : str
            The name of the sheet to update.
        cell : str
            The cell reference to update (e.g., 'A1').
        value : optional
            The value to set in the specified cell (default is 0).

        Raises
        ------
        ValueError
            If the cell reference or sheet name is invalid.
        TypeError
            If the sheet name is not a string.
        FileNotFoundError
            If the Excel file does not exist.
        """
        def valide_cell(cell):
            pattern = re.compile(r"^[A-Z]+[1-9]\d*$")
            if not pattern.match(cell):
                raise ValueError(f"Invalid cell reference: '{cell}'")

        try:
            valide_cell(cell)

            if sheetname is None:
                raise ValueError("Sheet name must be non-empty string.")

            if type(sheetname) is not str:
                raise TypeError("Sheet name must be string.")

            if cell is None:
                raise ValueError("Cell must be non-empty string.")

            wb = load_workbook(self.filename)

            if sheetname in wb.sheetnames:
                sheet = wb[sheetname]
            else:
                print(f"{sheetname} does not exist")

            sheet[cell] = value

            wb.save(self.filename)
            print(f"Updeted cell {cell} in {sheetname} with value '{value}'.")

        except FileNotFoundError:
            print(f"Error: The file '{self.filename}' does not exist.")

        except ValueError as e:
            print(f"Value error: {e}")

        except TypeError as e:
            print(f"Type error: {e}")

        except Exception as e:
            print(f"An unexcepted error occured: {e}")

    def excel_open(self):

        try:
            if not os.path.exists(self.filename):
                print(f"File {self.filename} does not exist. Creating...")
                data = []
                self.excel_create(data)

            wb = load_workbook(self.filename)

            return wb
        except Exception as e:
            print(f"An unexpected error occured: {e}")

    def sheet_create(self, sheetname: str = None) -> None:
        """
        Creates a new sheet in the Excel file.

        Parameters
        ----------
        sheetname : str
            The name of the sheet to create.

        Raises
        ------
        ValueError
            If the sheet name is not provided.
        TypeError
            If the sheet name is not a string.
        FileNotFoundError
            If the Excel file does not exist.
        """
        try:

            if sheetname is None:
                raise ValueError("Sheet name must be non-empty string.")

            if type(sheetname) is not str:
                raise TypeError("Sheet name must be string.")

            wb = load_workbook(self.filename)

            if sheetname not in wb.sheetname:

                wb.create_sheet(title=sheetname)

                wb.save(self.filename)
            else:
                print(f"Sheet {sheetname} already exist in excel file.")

        except FileNotFoundError:
            print(f"Error: The file '{self.filename}' does not exist.")

        except ValueError as e:
            print(f"Value error: {e}")

        except TypeError as e:
            print(f"Type error: {e}")

        except Exception as e:
            print(f"An unexcepted error occured: {e}")

        else:
            print(f"New sheet {sheetname} is created successfully.")

    def sheet_name_change(self,
                          sheetname: str = None,
                          newname: str = None) -> None:
        """
        Renames an existing sheet in the Excel file.

        Parameters
        ----------
        sheetname : str
            The current name of the sheet.
        newname : str
            The new name for the sheet.

        Raises
        ------
        ValueError
            If the sheet name parameters are not provided.
        TypeError
            If the sheet name parameters are not strings.
        FileNotFoundError
            If the Excel file does not exist.
        """
        try:

            if sheetname is None or newname is None:
                raise ValueError("Sheet parameters must be non-empty strings.")

            if type(sheetname) is not str or type(newname) is not str:
                raise TypeError("Sheet name parameters must be string.")

            wb = load_workbook(self.filename)

            if sheetname in wb.sheetnames:
                sheet = wb[sheetname]

                sheet.title = newname

                wb.save(self.filename)

                print("Sheet name changed.")
            else:
                print(f"{sheetname} does not exist in excel file.")

        except FileNotFoundError:
            print(f"Error: The file '{self.filename}' does not exist.")

        except TypeError as e:
            print(f"Type error: {e}")

        except ValueError as e:
            print(f"Value error: {e}")

        except Exception as e:
            print(f"An unexcepted error occured: {e}")

    def sheet_delete(self, sheetname: str = None) -> None:
        """
        Deletes a sheet from the Excel file.

        Parameters
        ----------
        sheetname : str
            The name of the sheet to delete.

        Raises
        ------
        ValueError
            If the sheet name is not provided.
        TypeError
            If the sheet name is not a string.
        FileNotFoundError
            If the Excel file does not exist.
        """
        try:

            if sheetname is None:
                raise ValueError("Sheet name must be non-empty string.")

            if type(sheetname) is not str:
                raise TypeError("Sheet name must be string.")

            wb = load_workbook(self.filename)

            if sheetname in wb.sheetnames:

                wb.remove(wb[sheetname])

                wb.save(self.filename)
                print(f"The sheet '{sheetname}' has been deleted.")
            else:
                print(f"The sheet '{sheetname}' does not exist.")

        except FileNotFoundError:
            print(f"Error: The file '{self.filename}' does not exist.")

        except TypeError as e:
            print(f"Type error: {e}")

        except ValueError as e:
            print(f"Value error: {e}")

        except Exception as e:
            print(f"An unexcepted error occured: {e}")
